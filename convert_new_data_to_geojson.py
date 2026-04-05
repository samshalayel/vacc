import openpyxl
import json

# Load the Excel file
excel_file = "170120261350.xlsx"
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active

# Get column headers
headers = []
for cell in sheet[1]:
    headers.append(cell.value)

# Create GeoJSON structure
geojson = {
    "type": "FeatureCollection",
    "name": "location_point_unified_corrected_1",
    "crs": {
        "type": "name",
        "properties": {
            "name": "urn:ogc:def:crs:OGC:1.3:CRS84"
        }
    },
    "features": []
}

# Find x and y column indices
x_col_idx = headers.index("x") + 1 if "x" in headers else None
y_col_idx = headers.index("y") + 1 if "y" in headers else None

if not x_col_idx or not y_col_idx:
    print("Error: x or y columns not found!")
    exit(1)

# Get indices for all required fields
field_indices = {}
required_fields = [
    "Vaccination status of a Child | On Schedule",
    "Vaccination status of a Child | Defaulter",
    "Vaccination status of a Child | Zero Dose",
    "Total Children Vaccinated by Age | above 24",
    "Total Children Vaccinated by Age | 0 to 12",
    "Total Children Vaccinated by Age | 12 to 24",
    "Governorate",
    "Health Facility",
    "Suppervisor Name",
    "all_child"
]

for field in required_fields:
    if field in headers:
        field_indices[field] = headers.index(field) + 1

# Process each row
processed_count = 0
skipped_count = 0

for row_idx in range(2, sheet.max_row + 1):
    # Get coordinates
    x = sheet.cell(row=row_idx, column=x_col_idx).value
    y = sheet.cell(row=row_idx, column=y_col_idx).value

    # Skip if coordinates are missing
    if x is None or y is None:
        skipped_count += 1
        continue

    # Create feature properties
    properties = {}
    for field, col_idx in field_indices.items():
        value = sheet.cell(row=row_idx, column=col_idx).value
        properties[field] = value if value is not None else ""

    # Create feature
    feature = {
        "type": "Feature",
        "properties": properties,
        "geometry": {
            "type": "Point",
            "coordinates": [float(x), float(y)]
        }
    }

    geojson["features"].append(feature)
    processed_count += 1

# Save as JavaScript file
output_file = "data/location_point_unified_corrected_1.js"
with open(output_file, 'w', encoding='utf-8') as f:
    f.write("var json_location_point_unified_corrected_1 = ")
    json.dump(geojson, f, ensure_ascii=False)
    f.write(";")

print("=" * 80)
print("CONVERSION COMPLETE!")
print("=" * 80)
print(f"Total rows processed: {processed_count}")
print(f"Total rows skipped (no coordinates): {skipped_count}")
print(f"Output file: {output_file}")
print(f"Total features in GeoJSON: {len(geojson['features'])}")
print("\nGovernorate values in data:")
gov_values = set()
for feature in geojson['features']:
    if 'Governorate' in feature['properties']:
        gov = feature['properties']['Governorate']
        if gov:
            gov_values.add(gov)
for gov in sorted(gov_values):
    print(f"  - {gov}")
