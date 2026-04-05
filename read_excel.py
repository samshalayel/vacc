import openpyxl
import json

# Open Excel file
wb = openpyxl.load_workbook("S123_2adf56689a684dc2b08d6be6b905a2d6_EXCEL (22).xlsx")
sheet = wb.active

# Get headers
headers = []
for cell in sheet[1]:
    headers.append(cell.value)

print("Headers:")
for i, h in enumerate(headers):
    print(f"{i}: {h}")

print("\n\nFirst 3 rows of data:")
row_count = 0
for row in sheet.iter_rows(min_row=2, max_row=4, values_only=True):
    row_count += 1
    print(f"\nRow {row_count}:")
    for i, val in enumerate(row):
        if val is not None:
            print(f"  {headers[i]}: {val}")

print(f"\n\nTotal rows (including header): {sheet.max_row}")
