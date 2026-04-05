import openpyxl
import json

# Load the new Excel file
excel_file = "170120261350.xlsx"
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active

# Get column headers
headers = []
for cell in sheet[1]:
    headers.append(cell.value)

print("=" * 80)
print("EXCEL FILE STRUCTURE:")
print("=" * 80)
print(f"Total Columns: {len(headers)}")
print(f"Total Rows: {sheet.max_row}")

print("\n" + "=" * 80)
print("COLUMN HEADERS:")
print("=" * 80)
for i, header in enumerate(headers, 1):
    print(f"{i}. {header}")

print("\n" + "=" * 80)
print("REQUIRED FILTERS CHECK:")
print("=" * 80)
required_filters = [
    "Vaccination status of a Child | On Schedule",
    "Vaccination status of a Child | Defaulter",
    "Vaccination status of a Child | Zero Dose",
    "Total Children Vaccinated by Age | above 24",
    "Total Children Vaccinated by Age | 0 to 12",
    "Total Children Vaccinated by Age | 12 to 24",
    "Governorate",
    "Health Facility",
    "Suppervisor Name"
]

for filter_name in required_filters:
    if filter_name in headers:
        print(f"✓ {filter_name} - FOUND")
    else:
        print(f"✗ {filter_name} - NOT FOUND")

# Check for Governorate values
if "Governorate" in headers:
    gov_col_idx = headers.index("Governorate") + 1
    governorates = set()
    for row_idx in range(2, sheet.max_row + 1):
        gov_value = sheet.cell(row=row_idx, column=gov_col_idx).value
        if gov_value:
            governorates.add(gov_value)

    print("\n" + "=" * 80)
    print("GOVERNORATE VALUES:")
    print("=" * 80)
    for gov in sorted(governorates):
        print(f"  - {gov}")

# Display sample rows
print("\n" + "=" * 80)
print("FIRST 3 SAMPLE ROWS:")
print("=" * 80)
for row_idx in range(2, min(5, sheet.max_row + 1)):
    print(f"\nRow {row_idx - 1}:")
    for col_idx, header in enumerate(headers, 1):
        cell_value = sheet.cell(row=row_idx, column=col_idx).value
        if header in required_filters or header in ["latitude", "longitude", "Latitude", "Longitude"]:
            print(f"  {header}: {cell_value}")
    print("-" * 40)
