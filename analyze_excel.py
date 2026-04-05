import openpyxl
import json

# Load the Excel file
excel_file = r"C:\Users\Administrator\AppData\Local\Temp\qgis2web\qgis2web_2026_01_22-12_48_37_247602\data\S123_2adf56689a684dc2b08d6be6b905a2d6_EXCEL (23).xlsx"
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active

# Get column headers
headers = []
for cell in sheet[1]:
    headers.append(cell.value)

print("Excel File Structure:")
print("=" * 80)
print(f"Total Columns: {len(headers)}")
print(f"Total Rows: {sheet.max_row}")
print("\nColumn Headers:")
for i, header in enumerate(headers, 1):
    print(f"{i}. {header}")

print("\n" + "=" * 80)
print("\nFirst 5 Sample Rows:")
print("=" * 80)

# Display first 5 data rows
for row_idx in range(2, min(7, sheet.max_row + 1)):
    print(f"\nRow {row_idx - 1}:")
    for col_idx, header in enumerate(headers, 1):
        cell_value = sheet.cell(row=row_idx, column=col_idx).value
        print(f"  {header}: {cell_value}")
    print("-" * 80)

# Save headers to a file for reference
with open(r"C:\Users\Administrator\AppData\Local\Temp\qgis2web\qgis2web_2026_01_22-12_48_37_247602\excel_headers.json", 'w', encoding='utf-8') as f:
    json.dump(headers, f, ensure_ascii=False, indent=2)
