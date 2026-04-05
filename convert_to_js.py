import openpyxl
import json
from collections import defaultdict

# Load the Excel file
excel_file = r"C:\Users\Administrator\AppData\Local\Temp\qgis2web\qgis2web_2026_01_22-12_48_37_247602\data\S123_2adf56689a684dc2b08d6be6b905a2d6_EXCEL (23).xlsx"
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active

# Load existing location data to get coordinates
existing_js = r"C:\Users\Administrator\AppData\Local\Temp\qgis2web\qgis2web_2026_01_22-12_48_37_247602\data\location_point_unified_corrected_1.js"
with open(existing_js, 'r', encoding='utf-8') as f:
    content = f.read()
    # Extract JSON part
    json_start = content.find('{')
    json_str = content[json_start:]
    existing_data = json.loads(json_str)

# Create a mapping of facility names to coordinates
facility_coords = {}
for feature in existing_data['features']:
    name_en = feature['properties'].get('Medical Point - Health Facility Name in English', '')
    name_ar = feature['properties'].get('Medical Point - Health Facility Name in Arabic', '')
    coords = feature['geometry']['coordinates']
    if name_en:
        facility_coords[name_en] = coords

# Get column headers
headers = []
for cell in sheet[1]:
    headers.append(cell.value)

# Process Excel data and aggregate by facility
facility_data = defaultdict(lambda: {
    'facility_name': '',
    'governorate': '',
    'organization': '',
    'total_children': 0,
    'age_0_to_12': 0,
    'age_12_to_24': 0,
    'age_above_24': 0,
    'zero_dose': 0,
    'defaulter': 0,
    'on_schedule': 0,
    'coordinates': None,
    'vaccine_details': defaultdict(int),
    'muac_normal': 0,
    'muac_mam': 0,
    'muac_sam': 0,
    'muac_oedema': 0,
    'reports': []
})

# Process each row
for row_idx in range(2, sheet.max_row + 1):
    facility_name = sheet.cell(row=row_idx, column=headers.index('Health Facility') + 1).value
    if not facility_name:
        continue

    # Aggregate data for this facility
    data = facility_data[facility_name]
    data['facility_name'] = facility_name
    data['governorate'] = sheet.cell(row=row_idx, column=headers.index('Governorate') + 1).value or data['governorate']

    # Get report date
    report_date = sheet.cell(row=row_idx, column=headers.index('Report Date') + 1).value

    # Aggregate vaccination counts
    all_child = sheet.cell(row=row_idx, column=headers.index('all_child') + 1).value or 0
    data['total_children'] += int(all_child) if all_child else 0

    age_0_12 = sheet.cell(row=row_idx, column=headers.index('Total Children Vaccinated by Age | 0 to 12') + 1).value or 0
    data['age_0_to_12'] += int(age_0_12) if age_0_12 else 0

    age_12_24 = sheet.cell(row=row_idx, column=headers.index('Total Children Vaccinated by Age | 12 to 24') + 1).value or 0
    data['age_12_to_24'] += int(age_12_24) if age_12_24 else 0

    age_above_24 = sheet.cell(row=row_idx, column=headers.index('Total Children Vaccinated by Age | above 24') + 1).value or 0
    data['age_above_24'] += int(age_above_24) if age_above_24 else 0

    zero_dose = sheet.cell(row=row_idx, column=headers.index('Vaccination status of a Child | Zero Dose') + 1).value or 0
    data['zero_dose'] += int(zero_dose) if zero_dose else 0

    defaulter = sheet.cell(row=row_idx, column=headers.index('Vaccination status of a Child | Defaulter') + 1).value or 0
    data['defaulter'] += int(defaulter) if defaulter else 0

    on_schedule = sheet.cell(row=row_idx, column=headers.index('Vaccination status of a Child | On Schedule') + 1).value or 0
    data['on_schedule'] += int(on_schedule) if on_schedule else 0

    # MUAC screening (note: column names have HTML entities)
    muac_normal = sheet.cell(row=row_idx, column=headers.index('MUAC Screeninig (mm) | Normal &gt;= 125') + 1).value or 0
    data['muac_normal'] += int(muac_normal) if muac_normal else 0

    muac_mam = sheet.cell(row=row_idx, column=headers.index('MUAC Screeninig (mm) | MAM 115 - 124') + 1).value or 0
    data['muac_mam'] += int(muac_mam) if muac_mam else 0

    muac_sam = sheet.cell(row=row_idx, column=headers.index('MUAC Screeninig (mm) | SAM &lt;115') + 1).value or 0
    data['muac_sam'] += int(muac_sam) if muac_sam else 0

    muac_oedema = sheet.cell(row=row_idx, column=headers.index('MUAC Screenings (mm) | Oedema +, ++, +++') + 1).value or 0
    data['muac_oedema'] += int(muac_oedema) if muac_oedema else 0

    # Vaccine details
    vaccine_fields = ['Hep', 'BCG', 'IPV1', 'IPV2', 'Penta1', 'Penta2', 'Penta3',
                     'bOPV1', 'bOPV2', 'bOPV3', 'bOPV4', 'bOPV5',
                     'Rota1', 'Rota2', 'Rota3', 'PCV1', 'PCV2', 'PCV3',
                     'MMR1', 'MMR2', 'DTP', 'DT', 'Td']

    for vaccine in vaccine_fields:
        value = sheet.cell(row=row_idx, column=headers.index(vaccine) + 1).value or 0
        data['vaccine_details'][vaccine] += int(value) if value else 0

    # Store report info
    data['reports'].append({
        'date': str(report_date) if report_date else '',
        'supervisor': sheet.cell(row=row_idx, column=headers.index('Suppervisor Name') + 1).value or ''
    })

    # Get coordinates from existing data
    if facility_name in facility_coords:
        data['coordinates'] = facility_coords[facility_name]

# Create GeoJSON structure
features = []
facilities_without_coords = []

for facility_name, data in facility_data.items():
    if data['coordinates']:
        feature = {
            "type": "Feature",
            "properties": {
                "Medical Point - Health Facility Name": data['facility_name'],
                "Governorate": data['governorate'],
                "Total Children": data['total_children'],
                "Age 0-12 Months": data['age_0_to_12'],
                "Age 12-24 Months": data['age_12_to_24'],
                "Age Above 24 Months": data['age_above_24'],
                "Zero Dose": data['zero_dose'],
                "Defaulter": data['defaulter'],
                "On Schedule": data['on_schedule'],
                "MUAC Normal": data['muac_normal'],
                "MUAC MAM": data['muac_mam'],
                "MUAC SAM": data['muac_sam'],
                "MUAC Oedema": data['muac_oedema'],
                "Total Reports": len(data['reports']),
                "Vaccine Details": dict(data['vaccine_details'])
            },
            "geometry": {
                "type": "Point",
                "coordinates": data['coordinates']
            }
        }
        features.append(feature)
    else:
        facilities_without_coords.append(facility_name)

# Create final GeoJSON
geojson = {
    "type": "FeatureCollection",
    "name": "vaccination_data",
    "crs": {
        "type": "name",
        "properties": {
            "name": "urn:ogc:def:crs:OGC:1.3:CRS84"
        }
    },
    "features": features
}

# Write JavaScript file
output_file = r"C:\Users\Administrator\AppData\Local\Temp\qgis2web\qgis2web_2026_01_22-12_48_37_247602\data\vaccination_data.js"
with open(output_file, 'w', encoding='utf-8') as f:
    f.write("var json_vaccination_data = ")
    json.dump(geojson, f, ensure_ascii=False, indent=2)
    f.write(";")

# Print summary
print("\n" + "=" * 80)
print("CONVERSION SUMMARY")
print("=" * 80)
print(f"Total facilities processed: {len(facility_data)}")
print(f"Facilities with coordinates: {len(features)}")
print(f"Facilities without coordinates: {len(facilities_without_coords)}")
print(f"\nOutput file created: vaccination_data.js")
print(f"Total features in GeoJSON: {len(features)}")

if facilities_without_coords:
    print("\n" + "=" * 80)
    print("FACILITIES WITHOUT COORDINATES:")
    print("=" * 80)
    for facility in sorted(facilities_without_coords):
        print(f"  - {facility}")

print("\n" + "=" * 80)
print("SAMPLE FEATURE DATA:")
print("=" * 80)
if features:
    print(json.dumps(features[0], ensure_ascii=False, indent=2))
